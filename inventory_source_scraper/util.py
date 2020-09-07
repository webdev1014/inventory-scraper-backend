"""inventory_source_scraper Utility methods
"""
import os
import traceback
from openpyxl import Workbook, load_workbook
from pathlib import Path
from .database import Database
from math import ceil

output = os.path.join(os.path.dirname(__file__), 'static', 'output.xlsx')


def remove_output_file():
    file = Path(output)
    try:
        file.unlink()
    except FileNotFoundError:
        print('No output file')


def create_output_file():
    remove_output_file()

    database = Database()
    total_rows = database.get_num_rows()
    print('total_rows------', total_rows)
    batch_size = 100

    field_names = ['Product Name',
                   'UPC',
                   'Vendor',
                   'Company',
                   'Wholesale Price',
                   'MSRP',
                   'Amazon Price',
                   'Prime Shipping',
                   'Margin(%)',
                   'Margin($)']
    wb = Workbook()
    ws = wb.active
    for i, field_name in enumerate(field_names, start=1):
        ws.cell(row=1, column=i, value=field_name)
    wb.save(filename=output)

    for start_at in range(int(ceil(total_rows / batch_size * 1.0))):
        try:
            wb = load_workbook(filename=output)
            ws = wb.active
            row = ws.max_row + 1
            products = database.get_products(start_at, batch_size)

            for product in products:
                ws.cell(row=row, column=1, value=product['name'])
                ws.cell(row=row, column=2, value=product['upc'])
                ws.cell(row=row, column=3, value=product['vendor'])
                ws.cell(row=row, column=4, value=product['company'])
                ws.cell(row=row, column=5, value=product['price_inventory'])
                ws.cell(row=row, column=6, value=product['price_msrp'])
                ws.cell(row=row, column=7, value=product['price_amazon'])
                ws.cell(row=row, column=8, value=product['shipping_amazon'])
                ws.cell(row=row, column=9, value=f'=IF(G{row}="",(F{row}-E{row})/E{row}*100,(G{row}-E{row})/E{row}*100)')
                ws.cell(row=row, column=10, value=f'=IF(G{row}="",F{row}-E{row},G{row}-E{row})')
                row += 1
            print('start_at %s' % str(start_at))
            wb.save(filename=output)
        except:
            print(traceback.format_exc())
            pass


